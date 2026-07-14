"""文件服务 — workspace 文件操作、上传、搜索、上下文构建。

核心优化：WorkspaceCache 维护 .md 文件内容索引，避免每次请求全量 rglob + read。
- 首次调用：全量扫描目标目录，缓存 path/content/mtime
- 后续调用：只检查 mtime 变化的文件，增量更新
- search_workspace / build_context 均基于索引搜索，零 IO 开销
"""
import time
import base64
from pathlib import Path
from io import BytesIO
from threading import Lock
from app.config import (
    WORKSPACE, SCAN_IGNORE_DIRS,
    IMAGE_MAX_DIMENSION, IMAGE_JPEG_QUALITY,
)


class WorkspaceCache:
    """workspace .md 文件内容索引缓存。

    用法：
        cache = WorkspaceCache()
        context = cache.build_context(query)  # 搜索匹配的 .md 文件内容
        results = cache.search_workspace(query)  # 搜索匹配的文件路径列表

    缓存策略：
    - 馶次调用：全量扫描目标目录，缓存 {path: (content, mtime)}
    - 后续调用：遍历缓存中的文件，检查 mtime 是否变化
      - mtime 未变 → 使用缓存内容，零 IO
      - mtime 变化 → 重新读取该文件内容
    - 同时检查是否有新增文件（rglob），加入缓存
    - 被删除的文件从缓存中移除
    """

    def __init__(self):
        self._index: dict[str, tuple[str, float]] = {}  # rel_path → (content, mtime)
        self._lock = Lock()
        self._last_full_scan: float = 0  # 上次全量扫描时间戳
        self._full_scan_interval: float = 30  # 全量扫描间隔（秒），防遗漏新增文件

    def _get_target_dirs(self) -> list[Path]:
        """返回需要扫描的目录列表。"""
        return [WORKSPACE / d for d in ["memory", "projects", "tasks", "skill"]]

    def _refresh_index(self) -> None:
        """增量刷新文件索引。

        1. 检查缓存中每个文件的 mtime，只重新读取变更的文件
        2. 检查是否有新增文件（超过 _full_scan_interval 时做全量扫描）
        3. 移除已被删除的文件
        """
        target_dirs = self._get_target_dirs()
        now = time.time()
        need_full_scan = (now - self._last_full_scan) >= self._full_scan_interval

        if need_full_scan:
            # 全量扫描：找所有 .md 文件
            all_files: set[str] = set()
            for dir_path in target_dirs:
                if not dir_path.exists():
                    continue
                for f in dir_path.rglob("*.md"):
                    # 跳过忽略目录
                    if any(ignored in f.parts for ignored in SCAN_IGNORE_DIRS):
                        continue
                    rel_path = str(f.relative_to(WORKSPACE))
                    all_files.add(rel_path)

                    if rel_path in self._index:
                        # 已缓存 — 检查 mtime
                        cached_mtime = self._index[rel_path][1]
                        current_mtime = f.stat().st_mtime
                        if current_mtime != cached_mtime:
                            # 文件变更 → 重新读取
                            content = f.read_text(encoding="utf-8", errors="ignore")
                            self._index[rel_path] = (content, current_mtime)
                    else:
                        # 新文件 → 读取并缓存
                        content = f.read_text(encoding="utf-8", errors="ignore")
                        self._index[rel_path] = (content, f.stat().st_mtime)

            # 移除已删除的文件
            removed = set(self._index.keys()) - all_files
            for rel_path in removed:
                del self._index[rel_path]

            self._last_full_scan = now
        else:
            # 快速检查：只检查已缓存文件的 mtime
            changed_paths = []
            for rel_path, (content, cached_mtime) in list(self._index.items()):
                file_path = WORKSPACE / rel_path
                if not file_path.exists():
                    # 文件被删除
                    changed_paths.append(rel_path)
                    continue
                current_mtime = file_path.stat().st_mtime
                if current_mtime != cached_mtime:
                    # 文件变更 → 重新读取
                    new_content = file_path.read_text(encoding="utf-8", errors="ignore")
                    self._index[rel_path] = (new_content, current_mtime)

            for rel_path in changed_paths:
                del self._index[rel_path]

    def build_context(self, query: str) -> str:
        """构建知识库上下文 — 基于缓存索引搜索，避免全量文件 IO。

        搜索逻辑：在所有缓存的 .md 文件内容中做子串匹配，
        返回匹配文件的摘要（前 500 字符）。
        """
        with self._lock:
            self._refresh_index()

        results = []
        lower_query = query.lower()
        for rel_path, (content, _mtime) in self._index.items():
            if lower_query in content.lower():
                # 提取目录名（memory/projects/tasks/skill）
                parts = rel_path.replace("\\", "/").split("/")
                dir_name = parts[0] if len(parts) > 1 else "unknown"
                filename = Path(rel_path).name
                content_preview = content[:500] + "..." if len(content) > 500 else content
                results.append(f"【{dir_name}/{filename}】\n{content_preview}\n")

        return "\n".join(results)

    def search_workspace(self, query: str) -> dict:
        """搜索 workspace — 基于缓存索引，返回匹配文件列表。"""
        with self._lock:
            self._refresh_index()

        results = []
        lower_query = query.lower()
        for rel_path, (content, _mtime) in self._index.items():
            if lower_query in content.lower():
                parts = rel_path.replace("\\", "/").split("/")
                dir_name = parts[0] if len(parts) > 1 else "unknown"
                results.append({
                    "path": rel_path,
                    "type": dir_name,
                    "match": query
                })

        return {"query": query, "results": results}

    def stats(self) -> dict:
        """返回缓存统计信息（调试用）。"""
        return {
            "indexed_files": len(self._index),
            "last_full_scan": self._last_full_scan,
            "full_scan_interval": self._full_scan_interval,
        }


# 全局缓存实例 — module-level singleton，整个应用共享
_workspace_cache = WorkspaceCache()


def build_context(query: str) -> str:
    """构建知识库上下文（对外接口，委托 WorkspaceCache）。"""
    return _workspace_cache.build_context(query)


def search_workspace(query: str) -> dict:
    """搜索 workspace（对外接口，委托 WorkspaceCache）。"""
    return _workspace_cache.search_workspace(query)


# ──── 文件上传与内容提取（独立功能，不涉及缓存） ────


def _extract_docx_content(file_content: bytes) -> str:
    try:
        from docx import Document
        doc = Document(BytesIO(file_content))
        return "\n".join([paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip()])
    except Exception as e:
        return f"无法读取Word文档: {str(e)}"


def _extract_xlsx_content(file_content: bytes) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(file_content), read_only=True)
        content = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                row_str = "\t".join([str(cell) if cell else "" for cell in row])
                if row_str.strip():
                    rows.append(row_str)
            if rows:
                content.append(f"=== Sheet: {sheet_name} ===")
                content.extend(rows)
        wb.close()
        return "\n".join(content)
    except Exception as e:
        return f"无法读取Excel文档: {str(e)}"


def _extract_image_info(file_content: bytes, filename: str) -> str:
    try:
        from PIL import Image
        img = Image.open(BytesIO(file_content))
        info = f"图片: {filename}\n尺寸: {img.size}\n格式: {img.format}\n模式: {img.mode}"
        img.close()
        return info
    except Exception as e:
        return f"无法读取图片: {str(e)}"


# ──── 图片编码与压缩（多模态识别用） ────

_IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp")

_MIME_MAP = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".gif": "image/gif",
    ".bmp": "image/bmp",
    ".webp": "image/webp",
}


def is_image_file(filename: str) -> bool:
    """判断文件名是否为支持的图片类型。"""
    return Path(filename).suffix.lower() in _IMAGE_EXTENSIONS


def _compress_image(file_content: bytes, filename: str) -> tuple[bytes, str]:
    """压缩图片到合理尺寸，返回 (压缩后字节, mime类型)。

    策略：
    - GIF 保持原样（保留动图）
    - 其他格式：长边超过 IMAGE_MAX_DIMENSION 时缩放，统一转为 JPEG
    - 小于阈值的图片直接返回原数据
    """
    from PIL import Image

    ext = Path(filename).suffix.lower()

    # GIF 不压缩（保留动画帧）
    if ext == ".gif":
        return file_content, "image/gif"

    img = Image.open(BytesIO(file_content))

    # 处理 EXIF 旋转信息
    try:
        from PIL import ImageOps
        img = ImageOps.exif_transpose(img)
    except Exception:
        pass

    w, h = img.size
    needs_resize = max(w, h) > IMAGE_MAX_DIMENSION

    if needs_resize:
        ratio = IMAGE_MAX_DIMENSION / max(w, h)
        new_size = (int(w * ratio), int(h * ratio))
        img = img.resize(new_size, Image.LANCZOS)

    # 带 alpha 通道的先转 RGB（JPEG 不支持透明）
    if img.mode in ("RGBA", "LA", "P"):
        img = img.convert("RGB")
    elif img.mode != "RGB":
        img = img.convert("RGB")

    buf = BytesIO()
    img.save(buf, format="JPEG", quality=IMAGE_JPEG_QUALITY, optimize=True)
    img.close()
    return buf.getvalue(), "image/jpeg"


def encode_image_to_base64(file_content: bytes, filename: str) -> dict:
    """将图片压缩并编码为 base64 data URI 格式。

    返回: {"mime": "image/jpeg", "data": "base64字符串", "filename": "xxx.jpg", "original_size": 1024, "compressed_size": 512}
    """
    original_size = len(file_content)
    compressed_bytes, mime = _compress_image(file_content, filename)
    b64 = base64.b64encode(compressed_bytes).decode("utf-8")
    return {
        "mime": mime,
        "data": b64,
        "filename": filename,
        "original_size": original_size,
        "compressed_size": len(compressed_bytes),
    }


def extract_file_content(file_content: bytes, filename: str) -> str:
    lower_name = filename.lower()
    if lower_name.endswith(".docx"):
        return _extract_docx_content(file_content)
    elif lower_name.endswith(".xlsx"):
        return _extract_xlsx_content(file_content)
    elif lower_name.endswith((".jpg", ".jpeg", ".png", ".gif", ".bmp")):
        return _extract_image_info(file_content, filename)
    elif lower_name.endswith(".txt"):
        return file_content.decode("utf-8", errors="ignore")
    elif lower_name.endswith(".md"):
        return file_content.decode("utf-8", errors="ignore")
    else:
        return f"无法解析文件类型: {filename}"


def upload_file(file_content: bytes, filename: str):
    upload_dir = WORKSPACE / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    file_path = upload_dir / filename
    file_path.write_bytes(file_content)

    content = extract_file_content(file_content, filename)

    return {
        "status": "uploaded",
        "file": filename,
        "size": len(file_content),
        "content": content
    }
